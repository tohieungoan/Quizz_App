"""
API Endpoints for User management.
Includes role-based access control (RBAC) verification before execution.
"""
from typing import List, Any, Optional
from fastapi import APIRouter, Depends, HTTPException, status, BackgroundTasks, Query, UploadFile, File
import csv
import io
try:
    import openpyxl
except ImportError:
    openpyxl = None
from sqlalchemy.orm import Session
from pydantic import ValidationError

from app.api.deps import get_db, get_current_active_user, get_current_active_admin
from app.crud.crud_user import crud_user
from app.models.user import User, UserSetting
from app.models.exam import Exam, ExamAssignee
from app.models.group import Group, GroupMember
from app.schemas.user import (
    UserCreate,
    UserUpdate,
    UserResponse,
    UserSettingResponse,
    UserSettingUpdate,
    NotificationEmailRequest,
    UserImportRow,
    UserImportResult
)
from app.core.email import send_notification_email_verification
from app.core.config import settings
import secrets
import json
from app.core.redis import set_token
from app.db.session import SessionLocal
from app.models.notification import Notification
import uuid
import logging

logger = logging.getLogger(__name__)

router = APIRouter()

def process_import_background(valid_users: List[dict], admin_id: int, job_id: str):
    """
    Background task to securely hash passwords and insert users via PostgreSQL Upsert.
    """
    db = SessionLocal()
    try:
        # Import data using Upsert (ON CONFLICT DO NOTHING)
        imported_count = crud_user.bulk_create(db, users_in=valid_users)
        logger.info(f"Background import job ({job_id}) finished successfully. Imported {imported_count} users.")
        
        # Create Success Notification for the Admin
        notification = Notification(
            user_id=admin_id,
            title="Import Users Complete",
            content=f"Your background import job ({job_id}) has finished successfully. Imported {imported_count} users out of {len(valid_users)} valid records.",
            type="SYSTEM",
        )
        db.add(notification)
        db.commit()

        try:
            from app.api.v1.endpoints.exams import _send_sync_ws_notification
            _send_sync_ws_notification(
                user_id=admin_id,
                title="Import Users Complete",
                content=f"Your background import job ({job_id}) has finished successfully. Imported {imported_count} users out of {len(valid_users)} valid records."
            )
        except Exception:
            pass
    except Exception as e:
        logger.error(f"Error in background import job {job_id}: {str(e)}")
        # Notify Admin about failure
        notification = Notification(
            user_id=admin_id,
            title="Import Users Failed",
            content=f"Your background import job ({job_id}) failed due to an internal error: {str(e)}",
            type="SYSTEM",
        )
        db.add(notification)
        db.commit()

        try:
            from app.api.v1.endpoints.exams import _send_sync_ws_notification
            _send_sync_ws_notification(
                user_id=admin_id,
                title="Import Users Failed",
                content=f"Your background import job ({job_id}) failed due to an internal error: {str(e)}"
            )
        except Exception:
            pass
    finally:
        db.close()


@router.get("/", response_model=List[UserResponse], summary="Get list of users (Admin)")
def read_users(
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = Query(None, description="Search by email or full name"),
    role: Optional[str] = Query(None, description="Filter by user role (e.g. ADMIN, USER)"),
    status: Optional[str] = Query(None, description="Filter by account status (e.g. ACTIVE, LOCKED)"),
    current_user: User = Depends(get_current_active_admin),
):
    """
    Retrieve users list with pagination support (`skip`, `limit`) and filtering (`search`, `role`, `status`).
    **Required Permission**: Super Admin.
    """
    users = crud_user.get_multi(
        db, 
        skip=skip, 
        limit=limit,
        search=search,
        role=role,
        status=status
    )
    return users


@router.post("/", response_model=UserResponse, status_code=status.HTTP_201_CREATED, summary="Create new user (Admin)")
def create_user(
    user_in: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_admin),
):
    """
    Admin creates a new user in the system.
    **Required Permission**: Super Admin.
    """
    existing_user = crud_user.get_by_email(db, email=user_in.email)
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This email address is already registered.",
        )
    user = crud_user.create(db, obj_in=user_in)
    return user


@router.post("/import", response_model=UserImportResult, status_code=status.HTTP_202_ACCEPTED, summary="Import users from CSV/Excel (Admin)")
def import_users_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_admin),
):
    """
    Bulk import users via CSV or Excel file.
    **Required Permission**: Super Admin.
    Validates the structure and dispatches the heavy bcrypt hashing + insertion to a Background Task.
    Uses PostgreSQL Upsert to safely handle concurrent registrations without IntegrityError crashes.
    Max limit: 10,000 rows per file for enterprise safety.
    """
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="Only CSV or Excel files are allowed.")
    
    rows = []
    
    if file.filename.endswith('.csv'):
        # Use utf-8-sig to automatically strip the BOM character (\ufeff) that Excel adds when exporting CSV
        content = file.file.read().decode("utf-8-sig")
        csv_reader = csv.DictReader(io.StringIO(content))
        for row in csv_reader:
            # Strip leading/trailing whitespace from keys (column names) and values
            clean_row = {
                (k.strip() if k else k): (v.strip() if isinstance(v, str) else v) 
                for k, v in row.items()
            }
            rows.append(clean_row)
    else:
        if openpyxl is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel (.xlsx) support requires the openpyxl library. Please upload a .csv file."
            )
        wb = openpyxl.load_workbook(file.file, data_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]] if sheet.max_row > 0 else []
        for row_cells in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row_cells): # Skip entirely empty rows
                clean_cells = [str(cell).strip() if isinstance(cell, str) else cell for cell in row_cells]
                row_dict = dict(zip(headers, clean_cells))
                rows.append(row_dict)
    
    errors = []
    valid_users = []
    emails_in_file = set()
    row_count = 0
    
    for idx, row in enumerate(rows, start=2): # 1 is header
        row_count += 1
        if row_count > 10000:
            errors.append(f"Row {idx}: Limit exceeded. Maximum 10,000 users allowed per import.")
            break
            
        try:
            user_data = UserImportRow(**row)
            if user_data.email in emails_in_file:
                errors.append(f"Row {idx}: Duplicate email within the CSV file ({user_data.email}).")
            else:
                emails_in_file.add(user_data.email)
                valid_users.append(user_data.model_dump())
        except ValidationError as e:
            for err in e.errors():
                errors.append(f"Row {idx}: Field '{err['loc'][0]}' - {err['msg']}")
                
    if not errors and valid_users:
        # Pre-check existing emails (only those existing at this very moment) to give immediate validation feedback
        existing_emails = crud_user.check_existing_emails(db, list(emails_in_file))
        for idx, u in enumerate(valid_users, start=2):
            if u['email'] in existing_emails:
                errors.append(f"Row {idx}: Email already exists in database ({u['email']}).")

    if errors:
        return UserImportResult(
            success=False,
            message="Import failed due to validation errors. No users were imported.",
            imported_count=0,
            errors=errors
        )
        
    if not valid_users:
        return UserImportResult(success=False, message="The CSV file is empty.", imported_count=0, errors=[])
        
    # Dispatch Background Task
    job_id = str(uuid.uuid4())
    background_tasks.add_task(process_import_background, valid_users, current_user.id, job_id)
    
    return UserImportResult(
        success=True,
        message=f"File passed validation. {len(valid_users)} users are now being imported in the background. You will receive a notification when it completes.",
        job_id=job_id,
        imported_count=len(valid_users),
        errors=[]
    )


@router.get("/{user_id}", response_model=UserResponse, summary="Get user details")
def read_user_by_id(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Retrieve detailed user profile information by ID.
    **Required Permission**: Account owner OR Super Admin.
    """
    if current_user.id != user_id and current_user.role != "SUPER_ADMIN":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to view other users' information.",
        )

    user = crud_user.get_by_id(db, user_id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"User not found with ID {user_id}",
        )
    return user


@router.put("/{user_id}", response_model=UserResponse, summary="Update user information")
def update_user(
    user_id: int,
    user_in: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    Update user information by ID.
    **Required Permission**: Account owner OR Super Admin.
    *(Regular users are not allowed to update their own role, status, or email verification status)*
    """
    if current_user.id != user_id and current_user.role != "SUPER_ADMIN":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to edit other users' information.",
        )

    # Prevent regular users from updating administrative fields: role, status, or email_verified
    if current_user.role != "SUPER_ADMIN":
        forbidden_fields = {"role", "status", "email_verified"}
        intersect = forbidden_fields.intersection(user_in.model_fields_set)
        if intersect:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"You do not have permission to modify the following fields: {', '.join(intersect)}",
            )

    user = crud_user.get_by_id(db, user_id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"User not found with ID {user_id}",
        )
    updated_user = crud_user.update(db, db_obj=user, obj_in=user_in)
    return updated_user


@router.delete("/{user_id}", response_model=UserResponse, summary="Delete user (Admin)")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_admin),
):
    """
    Permanently delete a user by ID.
    **Required Permission**: Super Admin.
    """
    user = crud_user.get_by_id(db, user_id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"User not found with ID {user_id}",
        )
    deleted_user = crud_user.delete(db, user_id=user_id)
    return deleted_user


@router.get("/me/settings", response_model=UserSettingResponse, summary="Get current user's settings")
def get_user_settings(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Get settings for currently logged-in user.
    If settings don't exist, create default ones.
    """
    settings_obj = db.query(UserSetting).filter(UserSetting.user_id == current_user.id).first()
    if not settings_obj:
        settings_obj = UserSetting(user_id=current_user.id)
        db.add(settings_obj)
        db.commit()
        db.refresh(settings_obj)
    return settings_obj


@router.put("/me/settings", response_model=UserSettingResponse, summary="Update user notification settings")
def update_user_settings(
    settings_in: UserSettingUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Update notification preference flags for currently logged-in user.
    Does not allow direct modification of notification_email.
    """
    settings_obj = db.query(UserSetting).filter(UserSetting.user_id == current_user.id).first()
    if not settings_obj:
        settings_obj = UserSetting(user_id=current_user.id)
        db.add(settings_obj)
        db.commit()
        db.refresh(settings_obj)

    update_data = settings_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if hasattr(settings_obj, field):
            setattr(settings_obj, field, value)

    db.add(settings_obj)
    db.commit()
    db.refresh(settings_obj)
    return settings_obj


@router.post("/me/notification-email/request", summary="Request verification email for a new notification email")
async def request_notification_email_verification(
    body: NotificationEmailRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Send a verification token to the requested notification email address.
    Does not update the email address until user confirms via verification token link.
    """
    # Check if this email is already registered as notification email by another user
    existing_setting = db.query(UserSetting).filter(
        UserSetting.notification_email == body.email,
        UserSetting.user_id != current_user.id
    ).first()
    if existing_setting:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This email address is already in use by another account for notifications."
        )

    token = secrets.token_urlsafe(32)
    payload = json.dumps({"user_id": current_user.id, "new_email": body.email})
    await set_token(f"verify_notification_email:{token}", payload, expire_seconds=7200)

    verify_url = f"{settings.FRONTEND_URL}/verify-notification-email?token={token}"
    background_tasks.add_task(
        send_notification_email_verification,
        email_to=body.email,
        verify_url=verify_url
    )
    return {"message": "Verification email has been sent successfully."}


@router.get("/{user_id}/assigned-exams", summary="Get all exams assigned to a specific user")
def get_user_assigned_exams(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Retrieve all exams assigned to a specific user (via direct assignment or group membership).
    """
    if current_user.id != user_id and current_user.role not in ["SUPER_ADMIN", "ADMIN", "HOST", "TEACHER"]:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to view this user's exams.",
        )

    user = crud_user.get_by_id(db, user_id=user_id)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"User not found with ID {user_id}",
        )

    assignees = db.query(ExamAssignee).filter(ExamAssignee.user_id == user_id).all()
    result = []
    for assignee in assignees:
        exam = assignee.exam
        if not exam:
            continue
        result.append({
            "id": exam.id,
            "assignee_id": assignee.id,
            "title": exam.title,
            "subject": exam.quiz.subject if (exam.quiz and exam.quiz.subject) else (exam.group.name if exam.group else "General"),
            "status": assignee.status,
            "score": assignee.score,
            "timer": exam.timer,
            "start_time": exam.start_time,
            "end_time": exam.end_time,
            "group_name": exam.group.name if exam.group else None,
            "quiz_title": exam.quiz.title if exam.quiz else None,
            "submitted_at": assignee.submitted_at,
        })
    return result