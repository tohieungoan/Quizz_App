from sqlalchemy.orm import Session, selectinload
from sqlalchemy import func, desc, or_, case, literal, and_
import io
import zipfile
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.quiz import Quiz, Question
from app.models.user import User
from app.models.room import Room, Participant, ParticipantAnswer
from app.models.exam import Exam, ExamAssignee, ExamAnswer
from app.models.quiz_variant import QuizVariant, QuizVariantQuestion
from app.schemas.report import ReportMetrics, ReportListItem, ReportPageResponse, ReportParticipant, ReportQuestionAnalysis, ReportParticipantPageResponse, ReportQuestionPageResponse

class CRUDReport:
    def get_metrics(self, db: Session) -> ReportMetrics:
        total_participants = db.query(func.count(Participant.id)).scalar() or 0
        total_assignees = db.query(func.count(ExamAssignee.id)).scalar() or 0
        total_users = total_participants + total_assignees

        total_questions = db.query(func.count(Question.id)).scalar() or 0

        # Calculate average score percentage from ParticipantAnswer
        # Assuming is_correct is boolean
        total_answers = db.query(func.count(ParticipantAnswer.id)).scalar() or 0
        correct_answers = db.query(func.count(ParticipantAnswer.id)).filter(ParticipantAnswer.is_correct == True).scalar() or 0
        
        avg_score = 0.0
        if total_answers > 0:
            avg_score = (correct_answers / total_answers) * 100.0

        return ReportMetrics(
            avg_score=round(avg_score, 1),
            total_participants=total_users,
            total_questions=total_questions
        )

    def get_reports_paginated(
        self, db: Session, search: str = "", report_type: str = "ALL", skip: int = 0, limit: int = 5
    ) -> ReportPageResponse:
        normalized_type = report_type.upper()
        search_term = f"%{search.strip()}%"

        # Aggregate child tables once and join the summaries into their parent
        # sessions. This avoids running three extra queries for every report.
        room_participant_counts = (
            db.query(
                Participant.room_id.label("room_id"),
                func.count(Participant.id).label("participant_count"),
            )
            .group_by(Participant.room_id)
            .subquery()
        )
        room_answer_stats = (
            db.query(
                Participant.room_id.label("room_id"),
                func.count(ParticipantAnswer.id).label("total_answers"),
                func.sum(
                    case((ParticipantAnswer.is_correct.is_(True), 1), else_=0)
                ).label("correct_answers"),
            )
            .outerjoin(
                ParticipantAnswer,
                ParticipantAnswer.participant_id == Participant.id,
            )
            .group_by(Participant.room_id)
            .subquery()
        )
        exam_assignee_counts = (
            db.query(
                ExamAssignee.exam_id.label("exam_id"),
                func.count(ExamAssignee.id).label("participant_count"),
            )
            .group_by(ExamAssignee.exam_id)
            .subquery()
        )
        exam_answer_stats = (
            db.query(
                ExamAssignee.exam_id.label("exam_id"),
                func.count(ExamAnswer.id).label("total_answers"),
                func.sum(
                    case((ExamAnswer.is_correct.is_(True), 1), else_=0)
                ).label("correct_answers"),
            )
            .outerjoin(
                ExamAnswer,
                ExamAnswer.exam_assignee_id == ExamAssignee.id,
            )
            .group_by(ExamAssignee.exam_id)
            .subquery()
        )

        report_queries = []

        if normalized_type in ["ALL", "ROOM"]:
            room_query = (
                db.query(
                    Room.id.label("id"),
                    literal("ROOM").label("type"),
                    func.coalesce(Room.room_code, "").label("room_code"),
                    func.coalesce(Quiz.title, "").label("quiz_title"),
                    func.coalesce(Room.title, Room.room_code, "").label("room_title"),
                    func.coalesce(User.fullname, "").label("host"),
                    func.coalesce(Room.ended_at, Room.created_at).label("event_date"),
                    func.coalesce(room_participant_counts.c.participant_count, 0).label("participants"),
                    func.coalesce(room_answer_stats.c.total_answers, 0).label("total_answers"),
                    func.coalesce(room_answer_stats.c.correct_answers, 0).label("correct_answers"),
                )
                .join(Quiz, Room.quiz_id == Quiz.id)
                .join(User, Room.host_id == User.id)
                .outerjoin(
                    room_participant_counts,
                    room_participant_counts.c.room_id == Room.id,
                )
                .outerjoin(
                    room_answer_stats,
                    room_answer_stats.c.room_id == Room.id,
                )
                .filter(Room.status == "ENDED")
            )
            if search.strip():
                room_query = room_query.filter(
                    or_(
                        Quiz.title.ilike(search_term),
                        Room.title.ilike(search_term),
                        Room.room_code.ilike(search_term),
                        User.fullname.ilike(search_term),
                    )
                )
            report_queries.append(room_query)

        if normalized_type in ["ALL", "EXAM"]:
            exam_query = (
                db.query(
                    Exam.id.label("id"),
                    literal("EXAM").label("type"),
                    literal("").label("room_code"),
                    func.coalesce(Quiz.title, "").label("quiz_title"),
                    func.coalesce(Exam.title, "").label("room_title"),
                    func.coalesce(User.fullname, "").label("host"),
                    func.coalesce(Exam.end_time, Exam.created_at).label("event_date"),
                    func.coalesce(exam_assignee_counts.c.participant_count, 0).label("participants"),
                    func.coalesce(exam_answer_stats.c.total_answers, 0).label("total_answers"),
                    func.coalesce(exam_answer_stats.c.correct_answers, 0).label("correct_answers"),
                )
                .join(Quiz, Exam.quiz_id == Quiz.id)
                .join(User, Exam.host_id == User.id)
                .outerjoin(
                    exam_assignee_counts,
                    exam_assignee_counts.c.exam_id == Exam.id,
                )
                .outerjoin(
                    exam_answer_stats,
                    exam_answer_stats.c.exam_id == Exam.id,
                )
                .filter(Exam.status.in_(["ACTIVE", "ENDED", "FINISHED", "COMPLETED"]))
            )
            if search.strip():
                exam_query = exam_query.filter(
                    or_(
                        Quiz.title.ilike(search_term),
                        Exam.title.ilike(search_term),
                        User.fullname.ilike(search_term),
                    )
                )
            report_queries.append(exam_query)

        if not report_queries:
            return ReportPageResponse(
                data=[],
                total=0,
                pageIndex=(skip // limit) + 1 if limit > 0 else 1,
                pageSize=limit,
            )

        combined_query = report_queries[0]
        if len(report_queries) > 1:
            combined_query = combined_query.union_all(*report_queries[1:])
        combined_reports = combined_query.subquery()

        # Count and paginate in SQL. Only the requested rows are materialized.
        total = db.query(func.count()).select_from(combined_reports).scalar() or 0
        rows = (
            db.query(combined_reports)
            .order_by(
                combined_reports.c.event_date.desc(),
                combined_reports.c.type.asc(),
                combined_reports.c.id.desc(),
            )
            .offset(skip)
            .limit(limit)
            .all()
        )

        paginated_reports = []
        for row in rows:
            total_answers = int(row.total_answers or 0)
            correct_answers = int(row.correct_answers or 0)
            avg_score = (
                f"{(correct_answers / total_answers * 100):.1f}%"
                if total_answers > 0
                else "0.0%"
            )
            event_date = row.event_date.strftime("%Y-%m-%d %H:%M") if row.event_date else ""
            room_code = row.room_code or (f"EX-{row.id}" if row.type == "EXAM" else "")
            room_title = row.room_title or (f"Exam {row.id}" if row.type == "EXAM" else room_code)

            paginated_reports.append(
                ReportListItem(
                    id=row.id,
                    type=row.type,
                    room_code=room_code,
                    quiz_title=row.quiz_title or "",
                    room_title=room_title,
                    host=row.host or "",
                    date=event_date,
                    participants=int(row.participants or 0),
                    avg_score=avg_score,
                )
            )

        return ReportPageResponse(
            data=paginated_reports,
            total=total,
            pageIndex=(skip // limit) + 1 if limit > 0 else 1,
            pageSize=limit
        )

    def get_report_participants(self, db: Session, session_id: int, session_type: str, skip: int = 0, limit: int = 50) -> ReportParticipantPageResponse:
        results = []
        total = 0
        
        if session_type.upper() == "ROOM":
            query = db.query(Participant).options(
                selectinload(Participant.quiz_variant)
            ).filter(Participant.room_id == session_id)
            total = query.count()
            # Get all participants sorted by score for ranking
            all_participants = query.order_by(desc(Participant.score)).all()
            
            # Build rank map (1-indexed)
            rank_map = {}
            for idx, p in enumerate(all_participants):
                rank_map[p.id] = idx + 1
            
            # Paginate
            paginated = all_participants[skip:skip + limit]
            
            for p in paginated:
                # Calculate correct/total answers
                total_ans = db.query(func.count(ParticipantAnswer.id)).filter(
                    ParticipantAnswer.participant_id == p.id
                ).scalar() or 0
                correct_ans = db.query(func.count(ParticipantAnswer.id)).filter(
                    ParticipantAnswer.participant_id == p.id,
                    ParticipantAnswer.is_correct == True
                ).scalar() or 0
                
                acc = f"{(correct_ans / total_ans * 100):.1f}%" if total_ans > 0 else "0%"
                
                results.append(ReportParticipant(
                    id=f"P-{p.id}",
                    user_id=str(p.user_id) if hasattr(p, 'user_id') and p.user_id else None,
                    nickname=p.nickname or "Anonymous",
                    status=p.status or "Completed",
                    joined_at=p.joined_at.strftime("%Y-%m-%d %H:%M") if hasattr(p, 'joined_at') and p.joined_at else None,
                    score=p.score,
                    correct_answers=f"{correct_ans}/{total_ans}",
                    accuracy=acc,
                    rank=rank_map.get(p.id, 0),
                    version_code=p.quiz_variant.version_code if p.quiz_variant else None,
                ))
        elif session_type.upper() == "EXAM":
            query = db.query(ExamAssignee, User.fullname).options(
                selectinload(ExamAssignee.quiz_variant)
            ).outerjoin(User, ExamAssignee.user_id == User.id).filter(ExamAssignee.exam_id == session_id)
            total = query.count()
            # Get all for ranking
            all_assignees = query.order_by(desc(ExamAssignee.score)).all()
            
            # Build rank map
            rank_map = {}
            for idx, (a, _) in enumerate(all_assignees):
                rank_map[a.id] = idx + 1
            
            # Paginate
            paginated = all_assignees[skip:skip + limit]
            
            for a, fullname in paginated:
                u_name = fullname if fullname else "Anonymous"
                
                # Calculate correct/total answers
                total_ans = db.query(func.count(ExamAnswer.id)).filter(
                    ExamAnswer.exam_assignee_id == a.id
                ).scalar() or 0
                correct_ans = db.query(func.count(ExamAnswer.id)).filter(
                    ExamAnswer.exam_assignee_id == a.id,
                    ExamAnswer.is_correct == True
                ).scalar() or 0
                
                acc = f"{(correct_ans / total_ans * 100):.1f}%" if total_ans > 0 else "0%"
                
                results.append(ReportParticipant(
                    id=f"EA-{a.id}",
                    user_id=str(a.user_id) if hasattr(a, 'user_id') and a.user_id else None,
                    nickname=u_name,
                    status=a.status or "Completed",
                    joined_at=a.started_at.strftime("%Y-%m-%d %H:%M") if hasattr(a, 'started_at') and a.started_at else None,
                    score=a.score or 0,
                    correct_answers=f"{correct_ans}/{total_ans}",
                    accuracy=acc,
                    rank=rank_map.get(a.id, 0),
                    version_code=a.quiz_variant.version_code if a.quiz_variant else None,
                ))
                
        return ReportParticipantPageResponse(
            data=results,
            total=total,
            pageIndex=(skip // limit) + 1 if limit > 0 else 1,
            pageSize=limit
        )

    def get_report_questions(self, db: Session, session_id: int, session_type: str, skip: int = 0, limit: int = 50) -> ReportQuestionPageResponse:
        results: list[ReportQuestionAnalysis] = []
        total = 0
        normalized_type = session_type.upper()

        session = None
        answer_model = None
        answer_owner_column = None
        owner_ids = None
        if normalized_type == "ROOM":
            session = db.query(Room).filter(Room.id == session_id).first()
            answer_model = ParticipantAnswer
            answer_owner_column = ParticipantAnswer.participant_id
            owner_ids = db.query(Participant.id).filter(Participant.room_id == session_id)
        elif normalized_type == "EXAM":
            session = db.query(Exam).filter(Exam.id == session_id).first()
            answer_model = ExamAnswer
            answer_owner_column = ExamAnswer.exam_assignee_id
            owner_ids = db.query(ExamAssignee.id).filter(ExamAssignee.exam_id == session_id)

        if session and answer_model is not None and owner_ids is not None:
            if session.variant_set_id:
                query = (
                    db.query(
                        QuizVariantQuestion.id,
                        QuizVariantQuestion.original_question_id,
                        QuizVariant.version_code,
                        QuizVariantQuestion.content,
                        QuizVariantQuestion.difficulty,
                        func.count(answer_model.id).label("total_ans"),
                        func.sum(
                            case((answer_model.is_correct.is_(True), 1), else_=0)
                        ).label("correct_ans"),
                    )
                    .join(
                        QuizVariant,
                        QuizVariantQuestion.quiz_variant_id == QuizVariant.id,
                    )
                    .outerjoin(
                        answer_model,
                        and_(
                            answer_model.variant_question_id == QuizVariantQuestion.id,
                            answer_owner_column.in_(owner_ids),
                        ),
                    )
                    .filter(QuizVariant.variant_set_id == session.variant_set_id)
                    .group_by(
                        QuizVariantQuestion.id,
                        QuizVariant.version_code,
                    )
                    .order_by(
                        QuizVariant.version_code.asc(),
                        QuizVariantQuestion.position.asc(),
                        QuizVariantQuestion.id.asc(),
                    )
                )
                total = (
                    db.query(func.count(QuizVariantQuestion.id))
                    .join(QuizVariant, QuizVariantQuestion.quiz_variant_id == QuizVariant.id)
                    .filter(QuizVariant.variant_set_id == session.variant_set_id)
                    .scalar()
                    or 0
                )
                stats = query.offset(skip).limit(limit).all()
                for row in stats:
                    q_id, original_id, version_code, content, difficulty, total_ans, correct_ans = row
                    total_ans = int(total_ans or 0)
                    correct_ans = int(correct_ans or 0)
                    results.append(ReportQuestionAnalysis(
                        id=q_id,
                        original_question_id=original_id,
                        version_code=version_code,
                        question=content or "No content",
                        correct=correct_ans,
                        incorrect=total_ans - correct_ans,
                        rate=round((correct_ans / total_ans * 100), 2) if total_ans else 0.0,
                        difficulty=difficulty or "Medium",
                    ))
            else:
                query = (
                    db.query(
                        Question.id,
                        Question.content,
                        Question.difficulty,
                        func.count(answer_model.id).label("total_ans"),
                        func.sum(
                            case((answer_model.is_correct.is_(True), 1), else_=0)
                        ).label("correct_ans"),
                    )
                    .outerjoin(
                        answer_model,
                        and_(
                            answer_model.question_id == Question.id,
                            answer_owner_column.in_(owner_ids),
                        ),
                    )
                    .filter(Question.quiz_id == session.quiz_id)
                    .group_by(Question.id)
                    .order_by(Question.position.asc(), Question.id.asc())
                )
                total = (
                    db.query(func.count(Question.id))
                    .filter(Question.quiz_id == session.quiz_id)
                    .scalar()
                    or 0
                )
                stats = query.offset(skip).limit(limit).all()
                for row in stats:
                    q_id, content, difficulty, total_ans, correct_ans = row
                    total_ans = int(total_ans or 0)
                    correct_ans = int(correct_ans or 0)
                    results.append(ReportQuestionAnalysis(
                        id=q_id,
                        original_question_id=q_id,
                        version_code=None,
                        question=content or "No content",
                        correct=correct_ans,
                        incorrect=total_ans - correct_ans,
                        rate=round((correct_ans / total_ans * 100), 2) if total_ans else 0.0,
                        difficulty=difficulty or "Medium",
                    ))

        return ReportQuestionPageResponse(
            data=results,
            total=total,
            pageIndex=(skip // limit) + 1 if limit > 0 else 1,
            pageSize=limit
        )

    def export_report_zip(self, db: Session, session_id: int, session_type: str) -> bytes:
        """Build a ZIP containing separate participant and question workbooks."""
        normalized_type = session_type.upper()
        if normalized_type == "ROOM":
            session = db.query(Room).filter(Room.id == session_id).first()
            title = (session.title or session.room_code or f"Room {session_id}") if session else ""
        elif normalized_type == "EXAM":
            session = db.query(Exam).filter(Exam.id == session_id).first()
            title = (session.title or f"Exam {session_id}") if session else ""
        else:
            raise ValueError("Report type must be ROOM or EXAM.")
        if not session:
            raise ValueError("Report session not found.")

        participant_report = self.get_report_participants(
            db, session_id=session_id, session_type=normalized_type, skip=0, limit=1_000_000
        )
        question_report = self.get_report_questions(
            db, session_id=session_id, session_type=normalized_type, skip=0, limit=1_000_000
        )

        def version_key(code: str | None) -> str:
            if code:
                return code
            return "Unassigned" if session.variant_set_id else "Original"

        questions_by_version = defaultdict(list)
        for question in question_report.data:
            questions_by_version[version_key(question.version_code)].append(question)

        version_codes = sorted(
            set(questions_by_version),
            key=lambda value: (value not in {"A", "B", "C", "D", "E"}, value),
        )
        if not version_codes:
            version_codes = ["Original"]

        navy = "1A0B82"
        light_blue = "E8E7F7"
        light_gray = "F3F4F6"
        green = "E7F6EC"
        white = "FFFFFF"
        muted = "64748B"
        thin_gray = Side(style="thin", color="D9DEE7")
        bottom_border = Border(bottom=thin_gray)

        def style_title(sheet, text: str, end_column: str = "H") -> None:
            sheet.merge_cells(f"A1:{end_column}1")
            cell = sheet["A1"]
            cell.value = text
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color=white, bold=True, size=16)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            sheet.row_dimensions[1].height = 30
            sheet.sheet_view.showGridLines = False

        def style_header(sheet, row: int, columns: int = 8) -> None:
            for cell in sheet[row][:columns]:
                cell.fill = PatternFill("solid", fgColor=light_blue)
                cell.font = Font(color=navy, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = bottom_border
            sheet.row_dimensions[row].height = 28

        def finish_question_sheet(sheet) -> None:
            widths = [14, 22, 48, 14, 18, 14, 18, 18]
            for index, width in enumerate(widths, start=1):
                sheet.column_dimensions[get_column_letter(index)].width = width
            sheet.freeze_panes = "A6"

        def accuracy_value(value: str | None) -> float:
            raw = (value or "0%").strip().rstrip("%")
            try:
                return float(raw) / 100
            except ValueError:
                return 0

        def joined_at_value(value: str | None):
            if not value:
                return "N/A"
            try:
                return datetime.strptime(value, "%Y-%m-%d %H:%M")
            except ValueError:
                return value

        participant_workbook = Workbook()
        all_participants = participant_workbook.active
        all_participants.title = "Participants"
        style_title(all_participants, f"{title} — Participant Statistics", end_column="I")
        all_participants["A3"] = "Session ID"
        all_participants["B3"] = session_id
        all_participants["D3"] = "Type"
        all_participants["E3"] = normalized_type
        all_participants["G3"] = "Participants"
        all_participants["H3"] = participant_report.total
        participant_headers = [
            "Rank", "User ID", "Participant", "Version", "Score",
            "Correct Answers", "Accuracy", "Joined At", "Status",
        ]
        for column, header in enumerate(participant_headers, start=1):
            all_participants.cell(5, column, header)
        style_header(all_participants, 5, 9)
        participant_row = 6
        for participant in participant_report.data:
            all_participants.append([
                participant.rank,
                participant.user_id or "Guest",
                participant.nickname,
                version_key(participant.version_code),
                participant.score,
                participant.correct_answers,
                accuracy_value(participant.accuracy),
                joined_at_value(participant.joined_at),
                participant.status,
            ])
            all_participants.cell(participant_row, 7).number_format = "0.0%"
            if isinstance(all_participants.cell(participant_row, 8).value, datetime):
                all_participants.cell(participant_row, 8).number_format = "yyyy-mm-dd hh:mm"
            for column in range(1, 10):
                all_participants.cell(participant_row, column).alignment = Alignment(
                    horizontal="left" if column in {3, 9} else "center",
                    vertical="center",
                )
            if participant_row % 2 == 0:
                for cell in all_participants[participant_row][:9]:
                    cell.fill = PatternFill("solid", fgColor=light_gray)
            participant_row += 1
        if participant_row == 6:
            all_participants.append([None, None, "No participant data for this session."])
            all_participants.cell(6, 3).font = Font(color=muted, italic=True)
        else:
            all_participants.auto_filter.ref = f"A5:I{participant_row - 1}"
        all_participants.freeze_panes = "A6"
        all_participants.sheet_view.showGridLines = False
        for column, width in enumerate([10, 14, 28, 14, 14, 18, 14, 20, 18], start=1):
            all_participants.column_dimensions[get_column_letter(column)].width = width

        question_workbook = Workbook()
        question_workbook.remove(question_workbook.active)
        for code in version_codes:
            sheet_name = f"Version {code}" if code in {"A", "B", "C", "D", "E"} else code
            sheet = question_workbook.create_sheet(title=sheet_name[:31])
            style_title(sheet, f"{title} — Question Statistics — {sheet_name}")
            sheet["A3"] = "Session ID"
            sheet["B3"] = session_id
            sheet["D3"] = "Type"
            sheet["E3"] = normalized_type
            sheet["G3"] = "Version"
            sheet["H3"] = code
            question_headers = [
                "Question ID", "Original Question ID", "Question Content", "Difficulty",
                "Total Answers", "Correct", "Incorrect", "Accuracy",
            ]
            for column, header in enumerate(question_headers, start=1):
                sheet.cell(5, column, header)
            style_header(sheet, 5, 8)
            row = 6

            for question in questions_by_version.get(code, []):
                total_answers = question.correct + question.incorrect
                sheet.append([
                    question.id,
                    question.original_question_id or question.id,
                    question.question,
                    question.difficulty,
                    total_answers,
                    question.correct,
                    question.incorrect,
                    question.rate / 100,
                ])
                sheet.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
                sheet.cell(row, 3).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True, indent=1
                )
                for column in range(4, 9):
                    sheet.cell(row, column).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                sheet.cell(row, 8).number_format = "0.0%"
                if question.rate >= 70:
                    sheet.cell(row, 8).fill = PatternFill("solid", fgColor=green)
                if row % 2 == 0:
                    for cell in sheet[row][:8]:
                        if cell.fill.fill_type is None:
                            cell.fill = PatternFill("solid", fgColor=light_gray)
                sheet.row_dimensions[row].height = 32
                row += 1

            if not questions_by_version.get(code):
                sheet.append([None, None, "No question data for this version."])
                sheet.cell(row, 3).font = Font(color=muted, italic=True)

            if questions_by_version.get(code):
                sheet.auto_filter.ref = f"A5:H{row - 1}"
            finish_question_sheet(sheet)

        participant_output = io.BytesIO()
        participant_workbook.save(participant_output)
        question_output = io.BytesIO()
        question_workbook.save(question_output)

        zip_output = io.BytesIO()
        base_name = f"{normalized_type}_{session_id}"
        with zipfile.ZipFile(zip_output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                f"{base_name}_Participants.xlsx", participant_output.getvalue()
            )
            archive.writestr(
                f"{base_name}_Questions.xlsx", question_output.getvalue()
            )
        return zip_output.getvalue()

crud_report = CRUDReport()
